from flask import Flask
from flask import Blueprint
from flask import request
import psycopg2
from flask import json
from flask import jsonify
import datetime
import random
from flask import current_app
import base64
from decimal import Decimal
from .globals import reply
import base64
import uuid
import pandas as pd
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.platypus.flowables import Spacer
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus.paragraph import Paragraph
from reportlab.lib.styles import ParagraphStyle

# Add GTK3 to PATH for WeasyPrint on Windows
import os
gtk_path = r"C:\Program Files\GTK3-Runtime Win64\bin"
if os.path.exists(gtk_path) and gtk_path not in os.environ.get('PATH', ''):
    os.environ['PATH'] = gtk_path + os.pathsep + os.environ.get('PATH', '')

from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from jinja2 import Template
import pytz
import re
from cassandra.cluster import Cluster, ExecutionProfile, EXEC_PROFILE_DEFAULT
from cassandra.auth import PlainTextAuthProvider
from cassandra import ConsistencyLevel
from cassandra.policies import TokenAwarePolicy, DCAwareRoundRobinPolicy
from cassandra.query import SimpleStatement


CASSANDRA_KEYSPACE = 'navas_iot_dbx'
CASSANDRA_CONTACT_POINTS = ['165.232.128.208']
CASSANDRA_PORT = 9042
CASSANDRA_USERNAME = 'cassandra'
CASSANDRA_PASSWORD = 'Sterile-Nectar-Unrevised-Undertone-Stagnate1'
CASSANDRA_LOCAL_DC = 'datacenter1'

_cassandra_cluster = None
_cassandra_session = None

def get_cassandra_session():
    global _cassandra_cluster, _cassandra_session
    if _cassandra_session and not _cassandra_session.is_shutdown:
        return _cassandra_session
    try:
        auth_provider = PlainTextAuthProvider(
            username=CASSANDRA_USERNAME,
            password=CASSANDRA_PASSWORD
        )
        profile = ExecutionProfile(
            load_balancing_policy=TokenAwarePolicy(DCAwareRoundRobinPolicy(local_dc=CASSANDRA_LOCAL_DC)),
            consistency_level=ConsistencyLevel.ONE
        )
        _cassandra_cluster = Cluster(
            contact_points=CASSANDRA_CONTACT_POINTS,
            port=CASSANDRA_PORT,
            auth_provider=auth_provider,
            protocol_version=4,
            execution_profiles={EXEC_PROFILE_DEFAULT: profile}
        )
        _cassandra_session = _cassandra_cluster.connect(CASSANDRA_KEYSPACE)
        print("Successfully connected to Cassandra cluster")
        return _cassandra_session
    except Exception as e:
        print(f"Error connecting to Cassandra: {e}")
        return None

data_handler_bp = Blueprint("DataHandler", __name__)

timezone = pytz.timezone('Africa/Kampala')

#get state data of a device
@data_handler_bp.route("/data-house/devices/<device_uid>/state/<state_name>/duration", methods=["GET"])
def StateFinder(device_uid, state_name):

    try:
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _Focus_Device = str(device_uid)
        _Focus_StateName = str(state_name).upper()

        if(len(_Focus_Device) >5) and (len(_Focus_StateName) > 2):

            with _dbconnect:
                with _dbconnect.cursor() as cursor:
                    cursor.execute("SELECT duration_of_state FROM dll_state_durations WHERE device_imei=%s AND state=%s AND state_status=%s", (_Focus_Device, _Focus_StateName, 'active',))

                    if(cursor.rowcount == 1):

                        _dataTunnel = cursor.fetchone()
                        _State_ValueFound = int(_dataTunnel[0])
                        _DurationSpent = None

                        if(_State_ValueFound < 60):
                            _DurationSpent = f"{_State_ValueFound} Minutes"
                        else:
                            hours = _State_ValueFound // 60
                            minutes = _State_ValueFound % 60
                            _DurationSpent = f"{hours}Hr{'s' if hours > 1 else ''}-{minutes}Min{'s' if minutes != 1 else ''}"

                        _ReplyObject = {
                            "duration": _DurationSpent,
                            "state": _Focus_StateName
                        }
                        return reply("success", 200, "Duration Found", _ReplyObject)
                    
                    elif(cursor.rowcount == 0):
                        return reply("error", 400, "No Active State Data Found", "")

        else:
            return reply("error", 400, "Some data Is Missing", "")

    except Exception as error:
        return reply("error", 500, str(error), "")
    
#trips by API DATA
@data_handler_bp.route("/data-house/reports/trips/data", methods=["POST"])
def TripsLoader():
    try:
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()
        _cassandra_session = get_cassandra_session()

        _ReportFormat = str(_payload_data['data']['report_format'])
        _ReportDevices = _payload_data['data']['report_devices']
        _StartReport_Date = str(_payload_data['data']['start_date'])
        _EndReport_Date = str(_payload_data['data']['end_date'])

        if len(_ReportFormat) > 0 and len(_ReportDevices) > 0 and len(_StartReport_Date) > 4 and len(_EndReport_Date) > 4:
            # Fetch device names from Cassandra in bulk
            device_names = {}
            try:
                cass_query = _cassandra_session.prepare("""
                    SELECT device_imei, device_name 
                    FROM dll_device_basic_data 
                    WHERE device_imei IN ?
                """)
                cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                for d in cass_rows:
                    device_names[d.device_imei] = d.device_name
            except Exception:
                # Fallback: fetch one by one
                cass_query = _cassandra_session.prepare("""
                    SELECT device_imei, device_name 
                    FROM dll_device_basic_data 
                    WHERE device_imei = ?
                """)
                for imei in _ReportDevices:
                    cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                    device_names[imei] = cass_row.device_name if cass_row else "Unknown"

            with _dbconnect:
                with _dbconnect.cursor() as cursor:
                    _TemporaryTripsData_Holder = []

                    for deviceNominatedImei in _ReportDevices:
                        _deviceName = device_names.get(deviceNominatedImei, "Unknown")
                        
                        if _deviceName != "Unknown":
                            cursor.execute("""
                                SELECT 
                                    trip_uid, 
                                    trip_date, 
                                    start_time, 
                                    end_time, 
                                    start_mileage, 
                                    end_mileage, 
                                    start_fuel_level, 
                                    end_fuel_level, 
                                    driver_id, 
                                    starting_location_point, 
                                    end_location_point, 
                                    start_gps_cordinates, 
                                    end_gps_cordinates
                                FROM 
                                    dll_trips_auditor
                                WHERE 
                                    trip_date BETWEEN %s AND %s AND device_imei=%s AND trip_status=%s
                                ORDER BY id DESC
                            """, (_StartReport_Date, _EndReport_Date, deviceNominatedImei, 'ended'))

                            if cursor.rowcount >= 1:
                                _Single_DeviceTripObject = {
                                    f"{_deviceName}": []
                                }

                                _TripsDataTunnel_Link = cursor.fetchall()
                                
                                for _TripsDataTunnel in _TripsDataTunnel_Link:
                                    
                                    if(str(_TripsDataTunnel[4]) != 'NoData') and (str(_TripsDataTunnel[5]) != 'NoData'):
                                        _MileageCovered = round(Decimal(_TripsDataTunnel[4]), 2) - round(Decimal(_TripsDataTunnel[5]), 2)
                                    else:
                                        _MileageCovered = 'NoData'

                                    _SingleTrip = {
                                        "trip_uid": _TripsDataTunnel[0],
                                        "trip_date": _TripsDataTunnel[1],
                                        "start_time": _TripsDataTunnel[2],
                                        "end_time": _TripsDataTunnel[3],
                                        "start_mileage": _TripsDataTunnel[4],
                                        "end_mileage": _TripsDataTunnel[5],
                                        "mileage_covered": str(_MileageCovered).lstrip('-'),
                                        "start_fuel_level": _TripsDataTunnel[6],
                                        "end_fuel_level": _TripsDataTunnel[7],
                                        "driver_id": _TripsDataTunnel[8],
                                        "start_location": _TripsDataTunnel[9],
                                        "end_location": _TripsDataTunnel[10],
                                        "start_gps_cords": _TripsDataTunnel[11],
                                        "end_gps_cords": _TripsDataTunnel[12]
                                    }
                                    _Single_DeviceTripObject[f"{_deviceName}"].append(_SingleTrip)

                                _TemporaryTripsData_Holder.append(_Single_DeviceTripObject)
                        else:
                            # Skip device if not found
                            continue

                    # Only return the response after processing all devices
                    return reply("success", 200, "Found Trips", _TemporaryTripsData_Holder)
        else:
            return reply("error", 400, "Some data is Missing", "")

    except Exception as error:
        return reply("error", 500, str(error), "")



#trips by excell file
@data_handler_bp.route("/data-house/reports/trips/excel", methods=["POST"])
def TripsLoader_ByExcel_File():
    try:
        # Database connection and input payload
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _cassandra_session: Session = get_cassandra_session()
        if not _cassandra_session:
            print("Failed to get Cassandra session. Cannot process data.")

        _ReportDevices = _payload_data['data']['report_devices']
        _StartReport_Date = str(_payload_data['data']['start_date'])
        _EndReport_Date = str(_payload_data['data']['end_date'])
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        if len(_ReportDevices) > 0 and len(_StartReport_Date) > 4 and len(_EndReport_Date) > 4:
            with _dbconnect:
                with _dbconnect.cursor() as cursor:
                    # Log request
                    cursor.execute("""
                        INSERT INTO dll_reports_downloadable_files 
                            (request_uid, file_path, report_caller, request_status, request_datestamp, report_type)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (str(_OriginRequestID), 'NO_DIR_PATH', _OriginUserID, 'in_process', str(_RequestTime_Stamp), 'trips',))

                    # Postgres trips
                    device_query_placeholder = ', '.join(['%s'] * len(_ReportDevices))
                    cursor.execute(f"""
                        SELECT 
                            t.device_imei, 
                            t.trip_date, 
                            t.start_time, 
                            t.end_time, 
                            t.start_mileage, 
                            t.end_mileage, 
                            t.start_fuel_level, 
                            t.end_fuel_level, 
                            t.driver_id, 
                            t.starting_location_point, 
                            t.end_location_point, 
                            t.start_gps_cordinates, 
                            t.end_gps_cordinates
                        FROM dll_trips_auditor t
                        WHERE t.trip_date BETWEEN %s AND %s 
                        AND t.device_imei IN ({device_query_placeholder})
                        AND t.trip_status = %s
                        ORDER BY t.device_imei, t.trip_date DESC
                    """, [_StartReport_Date, _EndReport_Date, *_ReportDevices, 'ended'])

                    rows = cursor.fetchall()
                    if not rows:
                        cursor.execute(
                            "UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s",
                            ('failed', _OriginRequestID,)
                        )
                        return reply("error", 404, "No trips found for the selected devices", "")

                    # Cassandra device names
                    device_names = {}
                    try:
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei IN ?
                        """)
                        cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                        for d in cass_rows:
                            device_names[d.device_imei] = d.device_name
                    except Exception:
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei = ?
                        """)
                        for imei in set([r[0] for r in rows]):
                            cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                            device_names[imei] = cass_row.device_name if cass_row else "Unknown"

                    # Organize trips by device
                    Excel_Data = {}
                    for row in rows:
                        device_imei = row[0]
                        device_name = device_names.get(device_imei, "Unknown")

                        if str(row[4]) != 'NoData' and str(row[5]) != 'NoData':
                            _MileageCovered = round(Decimal(row[4]), 2) - round(Decimal(row[5]), 2)
                        else:
                            _MileageCovered = "NoData"

                        trip_data = [
                            row[1], row[2], row[3], row[4], row[5],
                            str(_MileageCovered).lstrip("-"), row[6], row[7], row[8],
                            row[9], row[10], row[11], row[12]
                        ]
                        sheet_key = f"{device_name[:15]}_{device_imei[-4:]}"
                        Excel_Data.setdefault(sheet_key, []).append(trip_data)

                    # --- 🔹 Generate Excel ---
                    reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                    os.makedirs(reports_dir, exist_ok=True)
                    random_suffix = random.randint(100000, 999999)
                    FileName = f"trips_report_{random_suffix}.xlsx"
                    file_path = os.path.join(reports_dir, FileName)

                    wb = Workbook()
                    for sheet_name, trips in Excel_Data.items():
                        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                        headers = [
                            "Trip Date", "Start Time", "End Time", "Start Mileage",
                            "End Mileage", "Distance Moved", "Start Fuel", "End Fuel",
                            "Driver ID", "Start Location", "End Location", "Start GPS", "End GPS"
                        ]
                        ws.append(headers)
                        for trip in trips:
                            ws.append(trip)

                        # Auto column width
                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                            col_letter = get_column_letter(col[0].column)
                            ws.column_dimensions[col_letter].width = max_length + 2

                    # Remove default sheet if still exists
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        del wb["Sheet"]

                    wb.save(file_path)

                    # Update database with file path
                    PhysicalPath = current_app.config['base_url'] + f"{reports_dir}/{FileName}"
                    cursor.execute("""
                        UPDATE dll_reports_downloadable_files
                        SET file_path=%s, request_status='completed'
                        WHERE request_uid=%s
                    """, (str(PhysicalPath), str(_OriginRequestID)))

                    return reply("success", 200, "Trips Excel report generated successfully", {"file_url": PhysicalPath})

        else:
            return reply("error", 400, "Some data is Missing", "")

    except Exception as error:
        return reply("error", 500, str(error), "")


@data_handler_bp.route("/data-house/reports/trips/pdf", methods=["POST"])
def TripsLoader_ByPDF_File():
    try:
        # Database connection and input payload
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _cassandra_session = get_cassandra_session()
        if not _cassandra_session:
            print("Failed to get Cassandra session. Cannot process data.")

        _ReportDevices = _payload_data['data']['report_devices']
        _StartReport_Date = str(_payload_data['data']['start_date'])
        _EndReport_Date = str(_payload_data['data']['end_date'])
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        if len(_ReportDevices) > 0 and len(_StartReport_Date) > 4 and len(_EndReport_Date) > 4:
            with _dbconnect:
                with _dbconnect.cursor() as cursor:
                    # Log request into the reports table
                    cursor.execute("""
                        INSERT INTO dll_reports_downloadable_files 
                            (request_uid, file_path, report_caller, request_status, request_datestamp, report_type)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (str(_OriginRequestID), 'NO_DIR_PATH', _OriginUserID, 'in_process', str(_RequestTime_Stamp), 'trips',))

                    # Prepare trip data for all devices in a single query (Postgres only)
                    device_query_placeholder = ', '.join(['%s'] * len(_ReportDevices))
                    cursor.execute(f"""
                        SELECT 
                            t.device_imei, 
                            t.trip_date, 
                            t.start_time, 
                            t.end_time, 
                            t.start_mileage, 
                            t.end_mileage, 
                            t.start_fuel_level, 
                            t.end_fuel_level, 
                            t.driver_id, 
                            t.starting_location_point, 
                            t.end_location_point, 
                            t.start_gps_cordinates, 
                            t.end_gps_cordinates
                        FROM dll_trips_auditor t
                        WHERE t.trip_date BETWEEN %s AND %s 
                        AND t.device_imei IN ({device_query_placeholder})
                        AND t.trip_status = %s
                        ORDER BY t.device_imei, t.trip_date DESC
                    """, [_StartReport_Date, _EndReport_Date, *_ReportDevices, 'ended'])

                    # Process fetched trips
                    rows = cursor.fetchall()
                    if not rows:
                        cursor.execute(
                            "UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", 
                            ('failed', _OriginRequestID,)
                        )
                        return reply("error", 404, "No trips found for the selected devices", "")

                    # --- 🔹 Fetch device names from Cassandra in bulk ---
                    device_names = {}
                    try:
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei IN ?
                        """)
                        cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                        for d in cass_rows:
                            device_names[d.device_imei] = d.device_name
                    except Exception as e:
                        # fallback in case IN is not supported / too large
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei = ?
                        """)
                        for imei in set([r[0] for r in rows]):
                            cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                            device_names[imei] = cass_row.device_name if cass_row else "Unknown"

                    # Organize data by device
                    PDF_Data = {}
                    for row in rows:
                        device_imei = row[0]
                        device_name = device_names.get(device_imei, "Unknown")

                        # Calculate mileage
                        if str(row[4]) != 'NoData' and str(row[5]) != 'NoData':
                            _MileageCovered = round(Decimal(row[4]), 2) - round(Decimal(row[5]), 2)
                        else:
                            _MileageCovered = "NoData"

                        trip_data = [
                            row[1], row[2], row[3], row[4], row[5], str(_MileageCovered).lstrip("-"), row[6],
                            row[7], row[8], row[9], row[10], row[11], row[12]
                        ]
                        sheet_key = f"{device_name[:15]}_{device_imei[-4:]}"
                        PDF_Data.setdefault(sheet_key, []).append(trip_data)

                    # --- 🔹 Generate PDF ---
                    reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                    os.makedirs(reports_dir, exist_ok=True)
                    random_suffix = random.randint(100000, 999999)
                    FileName = f"trips_report_{random_suffix}.pdf"
                    file_path = os.path.join(reports_dir, FileName)

                    html_template = Template("""
                        <!DOCTYPE html>
                        <html lang="en">
                        <head>
                            <meta charset="UTF-8">
                            <style>
                                @page {
                                    size: A4 landscape;
                                    margin: 1cm;
                                }
                                body {
                                    font-family: Arial, sans-serif;
                                    font-size: 9px;
                                }
                                h1 { text-align: center; }
                                h2 { text-align: left; }
                                .table-container {
                                    max-width: 100%;
                                    overflow-x: auto;
                                    margin-bottom: 20px;
                                }
                                table {
                                    width: 100%;
                                    border-collapse: collapse;
                                    font-size: 9px;
                                }
                                th, td {
                                    padding: 6px;
                                    border: 1px solid #ddd;
                                    text-align: left;
                                }
                                th {
                                    background-color: #f2f2f2;
                                }
                                tr:nth-child(even) {
                                    background-color: #f9f9f9;
                                }
                            </style>
                        </head>
                        <body>
                            <h1>FMS Trips Report</h1>
                            <b><br>
                            <h2>Period: {{_StartReport_Date}} TO {{_EndReport_Date}}</h2>
                            <br><br>
                            {% for sheet_name, trips in data.items() %}
                                <h2>Trips Report For: {{ sheet_name }}</h2>
                                <div class="table-container">
                                    <table>
                                        <thead>
                                            <tr>
                                                <th>Trip Date</th>
                                                <th>Start Time</th>
                                                <th>End Time</th>
                                                <th>Start Mileage</th>
                                                <th>End Mileage</th>
                                                <th>Distance Moved</th>
                                                <th>Start Fuel</th>
                                                <th>End Fuel</th>
                                                <th>Driver ID</th>
                                                <th>Start Location</th>
                                                <th>End Location</th>
                                                <th>Start GPS</th>
                                                <th>End GPS</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            {% for trip in trips %}
                                                <tr>
                                                    {% for field in trip %}
                                                        <td>{{ field }}</td>
                                                    {% endfor %}
                                                </tr>
                                            {% endfor %}
                                        </tbody>
                                    </table>
                                </div>
                            {% endfor %}
                        </body>
                        </html>
                    """)

                    # Render HTML with data
                    rendered_html = html_template.render(
                        data=PDF_Data, 
                        _StartReport_Date=_StartReport_Date, 
                        _EndReport_Date=_EndReport_Date
                    )

                    # Generate PDF using WeasyPrint
                    HTML(string=rendered_html).write_pdf(file_path)

                    # Update file path in the database
                    PhysicalPath = current_app.config['base_url'] + f"{reports_dir}/{FileName}"
                    cursor.execute("""
                        UPDATE dll_reports_downloadable_files
                        SET file_path=%s, request_status='completed'
                        WHERE request_uid=%s
                    """, (str(PhysicalPath), str(_OriginRequestID)))

                    return reply("success", 200, "Trips report generated successfully", {"file_url": PhysicalPath})

        else:
            return reply("error", 400, "Some data is Missing", "")

    except Exception as error:
        return reply("error", 500, str(error), "")


def safe_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

#fuel report by excell file
@data_handler_bp.route("/data-house/reports/fuel/excel", methods=["POST"])
def FuelLevelReport_ByFile():
    try:
        # Database connection and input payload
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        # Cassandra session
        _cassandra_session: Session = get_cassandra_session()
        if not _cassandra_session:
            print("Failed to get Cassandra session. Cannot process data.")

        _ReportDevices = _payload_data['data']['report_devices']
        _StartReport_Date = str(_payload_data['data']['start_date'])
        _EndReport_Date = str(_payload_data['data']['end_date'])
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        if len(_ReportDevices) > 0 and len(_StartReport_Date) > 4 and len(_EndReport_Date) > 4:
            with _dbconnect:
                with _dbconnect.cursor() as cursor:
                    # Log request
                    cursor.execute("""
                        INSERT INTO dll_reports_downloadable_files 
                            (request_uid, file_path, report_caller, request_status, request_datestamp, report_type)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        str(_OriginRequestID), 'NO_DIR_PATH', _OriginUserID,
                        'in_process', str(_RequestTime_Stamp), 'fuel',
                    ))

                    # --- 🔹 Postgres trips ---
                    device_query_placeholder = ', '.join(['%s'] * len(_ReportDevices))
                    cursor.execute(f"""
                        SELECT 
                            t.device_imei, 
                            t.trip_date, 
                            t.start_time, 
                            t.end_time, 
                            t.start_fuel_level, 
                            t.end_fuel_level, 
                            t.starting_location_point, 
                            t.end_location_point, 
                            t.start_mileage, 
                            t.end_mileage
                        FROM dll_trips_auditor t
                        WHERE t.trip_date BETWEEN %s AND %s 
                        AND t.device_imei IN ({device_query_placeholder})
                        AND t.trip_status = %s
                        ORDER BY t.device_imei, t.trip_date DESC
                    """, [_StartReport_Date, _EndReport_Date, *_ReportDevices, 'ended'])

                    rows = cursor.fetchall()
                    if not rows:
                        cursor.execute(
                            "UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s",
                            ('failed', _OriginRequestID,)
                        )
                        return reply("error", 404, "No trips found for the selected devices", "")

                    # --- 🔹 Cassandra device names ---
                    device_names = {}
                    try:
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei IN ?
                        """)
                        cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                        for d in cass_rows:
                            device_names[d.device_imei] = d.device_name
                    except Exception:
                        cass_query = _cassandra_session.prepare("""
                            SELECT device_imei, device_name 
                            FROM dll_device_basic_data 
                            WHERE device_imei = ?
                        """)
                        for imei in set([r[0] for r in rows]):
                            cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                            device_names[imei] = cass_row.device_name if cass_row else "Unknown"

                    # --- 🔹 Organize report by device ---
                    Excel_Data = {}
                    for row in rows:
                        device_imei = row[0]
                        device_name = device_names.get(device_imei, "Unknown")

                        trip_date, start_time, end_time = row[1], row[2], row[3]
                        s_fuel, e_fuel = safe_float(row[4]), safe_float(row[5])
                        start_location, end_location = row[6], row[7]
                        start_mileage, end_mileage = row[8], row[9]

                        # Fuel difference
                        fuel_diff = (s_fuel - e_fuel) if (s_fuel is not None and e_fuel is not None) else "NoData"

                        # Distance moved
                        if str(start_mileage) != "NoData" and str(end_mileage) != "NoData":
                            _MileageCovered = round(Decimal(start_mileage), 2) - round(Decimal(end_mileage), 2)
                        else:
                            _MileageCovered = "NoData"

                        trip_data = [
                            trip_date, start_time, end_time,
                            s_fuel, e_fuel,
                            str(fuel_diff).lstrip("-"), str(_MileageCovered).lstrip("-"),
                            start_location, end_location
                        ]

                        sheet_key = f"{device_name[:15]}_{device_imei[-4:]}"
                        Excel_Data.setdefault(sheet_key, []).append(trip_data)

                    # --- 🔹 Generate Excel ---
                    reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                    os.makedirs(reports_dir, exist_ok=True)
                    random_suffix = random.randint(100000, 999999)
                    FileName = f"fuel_report_{random_suffix}.xlsx"
                    file_path = os.path.join(reports_dir, FileName)

                    wb = Workbook()
                    for sheet_name, trips in Excel_Data.items():
                        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                        headers = [
                            "Trip Date", "Start Time", "End Time",
                            "Start Fuel Level", "End Fuel Level",
                            "Fuel Difference", "Distance Moved",
                            "Start Location", "End Location"
                        ]
                        ws.append(headers)
                        for trip in trips:
                            ws.append(trip)

                        # Auto column width
                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                            col_letter = get_column_letter(col[0].column)
                            ws.column_dimensions[col_letter].width = max_length + 2

                    # Remove default sheet if still exists
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        del wb["Sheet"]

                    wb.save(file_path)

                    # --- 🔹 Update DB with file path ---
                    PhysicalPath = current_app.config['base_url'] + f"{reports_dir}/{FileName}"
                    cursor.execute("""
                        UPDATE dll_reports_downloadable_files
                        SET file_path=%s, request_status='completed'
                        WHERE request_uid=%s
                    """, (str(PhysicalPath), str(_OriginRequestID)))

                    return reply("success", 200, "Fuel Excel report generated successfully", {"file_url": PhysicalPath})

        else:
            return reply("error", 400, "Some data is Missing", "")

    except Exception as error:
        return reply("error", 500, str(error), "")


@data_handler_bp.route("/data-house/reports/fuel/pdf", methods=["POST"])
def FuelLevelReport_ByPDF():
    try:
        import pytz
        tz = pytz.timezone("Africa/Nairobi")

        # --- DB connections + payload ---
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _cassandra_session = get_cassandra_session()
        if not _cassandra_session:
            print("Failed to get Cassandra session. Cannot process data.")

        _payload_data = request.get_json()
        _ReportDevices = _payload_data['data']['report_devices']
        _StartReport_Date = str(_payload_data['data']['start_date'])
        _EndReport_Date = str(_payload_data['data']['end_date'])
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(tz).strftime("%d-%m-%Y %I:%M:%S %p")

        if not (_ReportDevices and len(_StartReport_Date) > 4 and len(_EndReport_Date) > 4):
            return reply("error", 400, "Some data is Missing", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files 
                        (request_uid, file_path, report_caller, request_status, request_datestamp, report_type)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (str(_OriginRequestID), 'NO_DIR_PATH', _OriginUserID, 'in_process', str(_RequestTime_Stamp), 'fuel'))

                # --- 🔹 Fetch trips from Postgres ---
                device_query_placeholder = ', '.join(['%s'] * len(_ReportDevices))
                cursor.execute(f"""
                    SELECT 
                        t.device_imei, 
                        t.trip_date, 
                        t.start_time, 
                        t.end_time, 
                        t.start_fuel_level, 
                        t.end_fuel_level, 
                        t.driver_id, 
                        t.starting_location_point, 
                        t.end_location_point
                    FROM dll_trips_auditor t
                    WHERE t.trip_date BETWEEN %s AND %s 
                    AND t.device_imei IN ({device_query_placeholder})
                    AND t.trip_status = %s
                    ORDER BY t.device_imei, t.trip_date DESC
                """, [_StartReport_Date, _EndReport_Date, *_ReportDevices, 'ended'])

                rows = cursor.fetchall()
                if not rows:
                    cursor.execute(
                        "UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", 
                        ('failed', _OriginRequestID,)
                    )
                    return reply("error", 404, "No trips found for the selected devices", "")

                # --- 🔹 Fetch device names from Cassandra ---
                device_names = {}
                try:
                    cass_query = _cassandra_session.prepare("""
                        SELECT device_imei, device_name 
                        FROM dll_device_basic_data 
                        WHERE device_imei IN ?
                    """)
                    cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                    for d in cass_rows:
                        device_names[d.device_imei] = d.device_name
                except Exception:
                    cass_query = _cassandra_session.prepare("""
                        SELECT device_imei, device_name 
                        FROM dll_device_basic_data 
                        WHERE device_imei = ?
                    """)
                    for imei in set([r[0] for r in rows]):
                        cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                        device_names[imei] = cass_row.device_name if cass_row else "Unknown"

                # --- 🔹 Organize data ---
                PDF_Data = {}
                for row in rows:
                    device_imei = row[0]
                    device_name = device_names.get(device_imei, "Unknown")

                    start_fuel = row[4]
                    end_fuel = row[5]

                    if str(start_fuel) != 'NoData' and str(end_fuel) != 'NoData':
                        fuel_used = round(Decimal(start_fuel), 2) - round(Decimal(end_fuel), 2)
                    else:
                        fuel_used = "NoData"

                    trip_data = [
                        row[1], row[2], row[3], start_fuel, end_fuel, str(fuel_used).lstrip("-"),
                        row[6], row[7], row[8]
                    ]
                    sheet_key = f"{device_name[:15]}_{device_imei[-4:]}"
                    PDF_Data.setdefault(sheet_key, []).append(trip_data)

                # --- 🔹 Generate PDF ---
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                FileName = f"fuel_report_{random_suffix}.pdf"
                file_path = os.path.join(reports_dir, FileName)

                html_template = Template("""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            @page { size: A4 landscape; margin: 1cm; }
                            body { font-family: Arial, sans-serif; font-size: 9px; }
                            h1 { text-align: center; }
                            h2 { text-align: left; }
                            .table-container { max-width: 100%; overflow-x: auto; margin-bottom: 20px; }
                            table { width: 100%; border-collapse: collapse; font-size: 9px; }
                            th, td { padding: 6px; border: 1px solid #ddd; text-align: left; }
                            th { background-color: #f2f2f2; }
                            tr:nth-child(even) { background-color: #f9f9f9; }
                        </style>
                    </head>
                    <body>
                        <h1>FMS Fuel Level Report</h1>
                        <b><br>
                        <h2>Period: {{_StartReport_Date}} TO {{_EndReport_Date}}</h2>
                        <br><br>
                        {% for sheet_name, trips in data.items() %}
                            <h2>Fuel Report For: {{ sheet_name }}</h2>
                            <div class="table-container">
                                <table>
                                    <thead>
                                        <tr>
                                            <th>Trip Date</th>
                                            <th>Start Time</th>
                                            <th>End Time</th>
                                            <th>Start Fuel</th>
                                            <th>End Fuel</th>
                                            <th>Fuel Used</th>
                                            <th>Driver ID</th>
                                            <th>Start Location</th>
                                            <th>End Location</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for trip in trips %}
                                            <tr>
                                                {% for field in trip %}
                                                    <td>{{ field }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        {% endfor %}
                    </body>
                    </html>
                """)

                rendered_html = html_template.render(
                    data=PDF_Data,
                    _StartReport_Date=_StartReport_Date,
                    _EndReport_Date=_EndReport_Date
                )

                HTML(string=rendered_html).write_pdf(file_path)

                # Update file path in DB
                PhysicalPath = current_app.config['base_url'] + f"{reports_dir}/{FileName}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path=%s, request_status='completed'
                    WHERE request_uid=%s
                """, (str(PhysicalPath), str(_OriginRequestID)))

                return reply("success", 200, "Fuel report generated successfully", {"file_url": PhysicalPath})

    except Exception as error:
        return reply("error", 500, str(error), "")


#night driving reports
@data_handler_bp.route("/data-house/reports/night-driving/excel", methods=["POST"])
def NightDrivingReport_ByExcell():
    try:
        # --- 🔹 Database connections ---
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        # Cassandra session
        _cassandra_session: Session = get_cassandra_session()
        if not _cassandra_session:
            print("Failed to get Cassandra session. Cannot process data.")

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # --- 🔹 Validate ---
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # --- 🔹 Log report request ---
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'night_driving'))

                Excel_Data = {}

                # --- 🔹 Fetch device names from Cassandra ---
                device_names = {}
                try:
                    cass_query = _cassandra_session.prepare("""
                        SELECT device_imei, device_name 
                        FROM dll_device_basic_data 
                        WHERE device_imei IN ?
                    """)
                    cass_rows = _cassandra_session.execute(cass_query, (_ReportDevices,))
                    for d in cass_rows:
                        device_names[d.device_imei] = d.device_name
                except Exception:
                    # fallback: fetch per device
                    cass_query = _cassandra_session.prepare("""
                        SELECT device_imei, device_name 
                        FROM dll_device_basic_data 
                        WHERE device_imei = ?
                    """)
                    for imei in _ReportDevices:
                        cass_row = _cassandra_session.execute(cass_query, (imei,)).one()
                        device_names[imei] = cass_row.device_name if cass_row else "Unknown"

                # --- 🔹 Process each device ---
                for device_imei in _ReportDevices:
                    device_name = device_names.get(device_imei, "Unknown")

                    # Fetch Geozone Triggered Events (Postgres)
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') 
                            BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('night_driving', device_imei, _StartDate, _EndDate))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for Excel
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        time_from_device = event[1]
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        dateLogged = event[4]

                        geozone_report_data.append({
                            "Date Loged": dateLogged,
                            "Location Point": location_logged,
                            "Time Violated": time_from_device,
                            "Event Triggered": event_trigger,
                            "GPS Coordinates": location_coordinates
                        })

                    # Add to structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"
                    Excel_Data[sheet_name] = geozone_report_data

                # --- 🔹 Check if data exists ---
                if not Excel_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s",
                                   ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Night Driving Data to Report", "")

                # --- 🔹 Create Excel ---
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"nightdriving_report_{random_suffix}.xlsx"
                file_path = os.path.join(reports_dir, file_name)

                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, data in Excel_Data.items():
                        df = pd.DataFrame(data)
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                        # Adjust column widths
                        worksheet = writer.sheets[sheet_name]
                        for i, col in enumerate(df.columns):
                            max_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_width)

                # --- 🔹 Update DB with file path ---
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Night Driving report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")


@data_handler_bp.route("/data-house/reports/night-driving/pdf", methods=["POST"])
def NightDrivingReport_ByPDF():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'night_driving'))

                # Initialize PDF data dictionary
                PDF_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('night_driving', device_imei, _StartDate, _EndDate,))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for PDF Export
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        time_violated = event[1]
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        date_logged = event[4]

                        # Append to report data
                        geozone_report_data.append([
                            date_logged, 
                            location_logged, 
                            time_violated, 
                            event_trigger, 
                            location_coordinates
                        ])

                    # Add data to PDF structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    PDF_Data[sheet_name] = geozone_report_data

                # Check if there is data to export
                if not PDF_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Night Driving Data to Report", "")

                # Create PDF File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"nightdriving_report_{random_suffix}.pdf"
                file_path = os.path.join(reports_dir, file_name)

                # HTML Template for PDF using WeasyPrint
                html_template = Template("""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            @page {
                                size: A4 landscape;
                                margin: 1cm;
                            }
                            body {
                                font-family: Arial, sans-serif;
                                font-size: 9px;
                            }
                            h1 {
                                text-align: center;
                            }
                            h2 {
                                text-align: left;
                            }
                            .table-container {
                                max-width: 100%;
                                overflow-x: auto;
                                margin-bottom: 20px;
                            }
                            table {
                                width: 100%;
                                border-collapse: collapse;
                                font-size: 9px;
                            }
                            th, td {
                                padding: 6px;
                                border: 1px solid #ddd;
                                text-align: left;
                            }
                            th {
                                background-color: #f2f2f2;
                            }
                            tr:nth-child(even) {
                                background-color: #f9f9f9;
                            }
                        </style>
                    </head>
                    <body>
                        <h1>FMS Night Driving Event Report</h1>
                        <br><br>
                        <h2>Report Period : {{StartDate}} To {{EndDate}}</h2>
                        <br><br>
                        {% for sheet_name, events in data.items() %}
                            <h2>Report For: {{ sheet_name }}</h2>
                            <div class="table-container">
                                <table>
                                    <thead>
                                        <tr>
                                            <th>Date Logged</th>
                                            <th>Location Point</th>
                                            <th>Time Violated</th>
                                            <th>Event Triggered</th>
                                            <th>GPS Coordinates</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for event in events %}
                                            <tr>
                                                {% for field in event %}
                                                    <td>{{ field }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        {% endfor %}
                    </body>
                    </html>
                """)

                # Render HTML with data
                rendered_html = html_template.render(data=PDF_Data, StartDate=_StartDate, EndDate=_EndDate)

                # Generate PDF using WeasyPrint
                HTML(string=rendered_html).write_pdf(file_path)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Night Driving report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")
  


#Get geozone reports
@data_handler_bp.route("/data-house/reports/geozone/excel", methods=["POST"])
def GezoneReport_ByExcell():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'geozone'))

                # Initialize Excel data dictionary
                Excel_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('geozone', device_imei, _StartDate, _EndDate,))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for Excel Export
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        geozone_id = event[1].replace("inside_", "")
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        dateLogged = event[4]

                        # Get Geozone Name
                        cursor.execute("SELECT geozone_name FROM dll_geozones WHERE geozone_uid = %s", (geozone_id,))
                        geozone_info = cursor.fetchone()
                        geozone_name = geozone_info[0] if geozone_info else 'geozone_deleted'

                        # Append to report data
                        geozone_report_data.append({
                            "Date Loged": dateLogged,
                            "Location Point": location_logged,
                            "Geozone Name": geozone_name,
                            "Event Triggered": event_trigger,
                            "GPS Coordinates": location_coordinates
                        })

                    # Add data to Excel structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    Excel_Data[sheet_name] = geozone_report_data

                # Check if there is data to export
                if not Excel_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Geozone Data to Report", "")

                # Create Excel File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"geozone_report_{random_suffix}.xlsx"
                file_path = os.path.join(reports_dir, file_name)

                # Export data to Excel using Pandas
                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, data in Excel_Data.items():
                        df = pd.DataFrame(data)
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                        # Adjust column widths
                        worksheet = writer.sheets[sheet_name]
                        for i, col in enumerate(df.columns):
                            max_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_width)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Geozone report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")


@data_handler_bp.route("/data-house/reports/geozone/pdf", methods=["POST"])
def GezoneReport_ByPDF():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'geozone'))

                # Initialize PDF data dictionary
                PDF_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('geozone', device_imei, _StartDate, _EndDate,))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for PDF Export
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        geozone_id = event[1].replace("inside_", "")
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        date_logged = event[4]

                        # Get Geozone Name
                        cursor.execute("SELECT geozone_name FROM dll_geozones WHERE geozone_uid = %s", (geozone_id,))
                        geozone_info = cursor.fetchone()
                        geozone_name = geozone_info[0] if geozone_info else 'geozone_deleted'

                        # Append to report data
                        geozone_report_data.append([
                            date_logged, 
                            location_logged, 
                            geozone_name, 
                            event_trigger, 
                            location_coordinates
                        ])

                    # Add data to PDF structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    PDF_Data[sheet_name] = geozone_report_data

                # Check if there is data to export
                if not PDF_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Geozone Data to Report", "")

                # Create PDF File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"geozone_report_{random_suffix}.pdf"
                file_path = os.path.join(reports_dir, file_name)

                # HTML Template for PDF using WeasyPrint
                html_template = Template("""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            @page {
                                size: A4 landscape;
                                margin: 1cm;
                            }
                            body {
                                font-family: Arial, sans-serif;
                                font-size: 9px;
                            }
                            h1 {
                                text-align: center;
                            }
                            h2 {
                                text-align: left;
                            }
                            .table-container {
                                max-width: 100%;
                                overflow-x: auto;
                                margin-bottom: 20px;
                            }
                            table {
                                width: 100%;
                                border-collapse: collapse;
                                font-size: 9px;
                            }
                            th, td {
                                padding: 6px;
                                border: 1px solid #ddd;
                                text-align: left;
                            }
                            th {
                                background-color: #f2f2f2;
                            }
                            tr:nth-child(even) {
                                background-color: #f9f9f9;
                            }
                        </style>
                    </head>
                    <body>
                        <h1>FMS Geozone Event Report</h1>
                        <br><br>
                        <h2>Report Period : {{StartDate}} To {{EndDate}}</h2>
                        <br><br>
                        {% for sheet_name, events in data.items() %}
                            <h2>Report For: {{ sheet_name }}</h2>
                            <div class="table-container">
                                <table>
                                    <thead>
                                        <tr>
                                            <th>Date Logged</th>
                                            <th>Location Point</th>
                                            <th>Geozone Name</th>
                                            <th>Event Triggered</th>
                                            <th>GPS Coordinates</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for event in events %}
                                            <tr>
                                                {% for field in event %}
                                                    <td>{{ field }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        {% endfor %}
                    </body>
                    </html>
                """)

                # Render HTML with data
                rendered_html = html_template.render(data=PDF_Data, StartDate=_StartDate, EndDate=_EndDate)

                # Generate PDF using WeasyPrint
                HTML(string=rendered_html).write_pdf(file_path)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Geozone report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")


#Get overspeeding reports
@data_handler_bp.route("/data-house/reports/overspeeding/excel", methods=["POST"])
def OverSpeedingReport_ByExcell():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'overspeeding'))

                # Initialize Excel data dictionary
                Excel_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('speed', device_imei, _StartDate, _EndDate,))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for Excel Export
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        moving_speed = str(event[1])+"KM/H"
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        dateLogged = event[4]

                        # Append to report data
                        geozone_report_data.append({
                            "Date Loged": dateLogged,
                            "Location Point": location_logged,
                            "Event Triggered": event_trigger,
                            "Moving Speed": moving_speed,
                            "GPS Coordinates": location_coordinates
                        })

                    # Add data to Excel structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    Excel_Data[sheet_name] = geozone_report_data

                # Check if there is data to export
                if not Excel_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Overspeeding Data to Report", "")

                # Create Excel File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"overspeeding_report_{random_suffix}.xlsx"
                file_path = os.path.join(reports_dir, file_name)

                # Export data to Excel using Pandas
                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, data in Excel_Data.items():
                        df = pd.DataFrame(data)
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                        # Adjust column widths
                        worksheet = writer.sheets[sheet_name]
                        for i, col in enumerate(df.columns):
                            max_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_width)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Overspeeding report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")


@data_handler_bp.route("/data-house/reports/overspeeding/pdf", methods=["POST"])
def OverSpeedingReport_ByPDF():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, 'overspeeding'))

                # Initialize PDF data dictionary
                PDF_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT 
                            location_logged, 
                            value_from_device, 
                            value_triggered, 
                            location_cordinates,
                            date_logged
                        FROM dll_triggered_events 
                        WHERE event_triggered = %s 
                        AND device_imei = %s 
                        AND TO_TIMESTAMP(date_logged, 'DD-MM-YYYY') BETWEEN TO_TIMESTAMP(%s, 'DD-MM-YYYY') AND TO_TIMESTAMP(%s, 'DD-MM-YYYY')
                    """, ('speed', device_imei, _StartDate, _EndDate,))

                    geozone_events = cursor.fetchall()
                    if not geozone_events:
                        continue

                    # Prepare Data for PDF Export
                    geozone_report_data = []
                    for event in geozone_events:
                        location_logged = event[0]
                        movingSpeed = str(event[1])+"KM/H"
                        event_trigger = event[2]
                        location_coordinates = event[3]
                        date_logged = event[4]

                        # Append to report data
                        geozone_report_data.append([
                            date_logged, 
                            location_logged, 
                            movingSpeed, 
                            event_trigger, 
                            location_coordinates
                        ])

                    # Add data to PDF structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    PDF_Data[sheet_name] = geozone_report_data

                # Check if there is data to export
                if not PDF_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No Overspeeding Data to Report", "")

                # Create PDF File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"overspeeding_report_{random_suffix}.pdf"
                file_path = os.path.join(reports_dir, file_name)

                # HTML Template for PDF using WeasyPrint
                html_template = Template("""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            @page {
                                size: A4 landscape;
                                margin: 1cm;
                            }
                            body {
                                font-family: Arial, sans-serif;
                                font-size: 9px;
                            }
                            h1 {
                                text-align: center;
                            }
                            h2 {
                                text-align: left;
                            }
                            .table-container {
                                max-width: 100%;
                                overflow-x: auto;
                                margin-bottom: 20px;
                            }
                            table {
                                width: 100%;
                                border-collapse: collapse;
                                font-size: 9px;
                            }
                            th, td {
                                padding: 6px;
                                border: 1px solid #ddd;
                                text-align: left;
                            }
                            th {
                                background-color: #f2f2f2;
                            }
                            tr:nth-child(even) {
                                background-color: #f9f9f9;
                            }
                        </style>
                    </head>
                    <body>
                        <h1>FMS Overspeeding Events Report</h1>
                        <br><br>
                        <h2>Report Period : {{StartDate}} To {{EndDate}}</h2>
                        <br><br>
                        {% for sheet_name, events in data.items() %}
                            <h2>Report For: {{ sheet_name }}</h2>
                            <div class="table-container">
                                <table>
                                    <thead>
                                        <tr>
                                            <th>Date Logged</th>
                                            <th>Location Point</th>
                                            <th>Moving Speed</th>
                                            <th>Event Triggered</th>
                                            <th>GPS Coordinates</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for event in events %}
                                            <tr>
                                                {% for field in event %}
                                                    <td>{{ field }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        {% endfor %}
                    </body>
                    </html>
                """)

                # Render HTML with data
                rendered_html = html_template.render(data=PDF_Data, StartDate=_StartDate, EndDate=_EndDate)

                # Generate PDF using WeasyPrint
                HTML(string=rendered_html).write_pdf(file_path)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, "Overspeeding report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")
    

#State Duration Reports
@data_handler_bp.route("/data-house/reports/state/excel", methods=["POST"])
def StateReports_Excell():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _ReportState = _payload_data['data']['report_state']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, _ReportState,))

                # Initialize Excel data dictionary
                Excel_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT start_time, duration_of_state, date_logged, end_time, start_location, 
                            end_location, start_location_cords, end_location_cords 
                        FROM dll_state_durations 
                        WHERE device_imei=%s 
                        AND state=%s 
                        AND end_time != 'Incoming' 
                        AND date_logged BETWEEN %s AND %s
                    """, (device_imei, _ReportState, _StartDate, _EndDate))


                    state_data = cursor.fetchall()
                    if not state_data:
                        continue

                    # Prepare Data for Excel Export
                    state_report_data = []
                    for state_info in state_data:
                        _state_time = state_info[0]
                        _state_duration = int(state_info[1])
                        _date_logged = state_info[2]
                        _state_endTime = state_info[3]
                        _state_start_location = state_info[4]
                        _state_end_location = state_info[5]
                        _state_start_cords = state_info[6]
                        _state_end_cords = state_info[7]

                        _DurationSpent = None

                        if(_state_duration < 60):
                            _DurationSpent = f"{_state_duration} Minutes"
                        else:
                            hours = _state_duration // 60
                            minutes = _state_duration % 60
                            _DurationSpent = f"{hours}Hr{'s' if hours > 1 else ''}-{minutes}Min{'s' if minutes != 1 else ''}"

                        # Append to report data
                        state_report_data.append({
                            "Date Loged": _date_logged,
                            "Start Time": _state_time,
                            "End Time": _state_endTime,
                            "Duration": _DurationSpent,
                            "Start Location": _state_start_location,
                            "End Location": _state_end_location,
                            "Starting GPS": _state_start_cords,
                            "End GPS": _state_end_cords
                        })

                    # Add data to Excel structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"  # Ensure sheet name length and uniqueness
                    Excel_Data[sheet_name] = state_report_data

                # Check if there is data to export
                if not Excel_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No State Data to Report", "")

                # Create Excel File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"{_ReportState}_state_report_{random_suffix}.xlsx"
                file_path = os.path.join(reports_dir, file_name)

                # Export data to Excel using Pandas
                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    for sheet_name, data in Excel_Data.items():
                        df = pd.DataFrame(data)
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                        # Adjust column widths
                        worksheet = writer.sheets[sheet_name]
                        for i, col in enumerate(df.columns):
                            max_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_width)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, f"{_ReportState} State report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")



@data_handler_bp.route("/data-house/reports/state/pdf", methods=["POST"])
def StateReports_ToPDF():
    try:
        # Establish Database connection
        _dbconnect = psycopg2.connect(current_app.config['db_link'])
        _payload_data = request.get_json()

        _ReportDevices = _payload_data['data']['report_devices']
        _StartDate = _payload_data['data']['start_date']
        _EndDate = _payload_data['data']['end_date']
        _ReportState = _payload_data['data']['report_state']
        _OriginRequestID = str(uuid.uuid4())
        _OriginUserID = str(_payload_data['data']['origin_user'])
        _RequestTime_Stamp = datetime.datetime.now(timezone).strftime("%d-%m-%Y %I:%M:%S %p")

        # Validate Input Data
        if not _ReportDevices:
            return reply("error", 400, "No devices specified for the report", "")

        with _dbconnect:
            with _dbconnect.cursor() as cursor:
                # Log report request
                cursor.execute("""
                    INSERT INTO dll_reports_downloadable_files (
                        request_uid, file_path, report_caller, request_status, request_datestamp, report_type
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (_OriginRequestID, 'NO_DIR_PATH', _OriginUserID, 'in_process', _RequestTime_Stamp, _ReportState))

                # Initialize PDF data dictionary
                PDF_Data = {}

                # Process each device in the report request
                for device_imei in _ReportDevices:
                    cursor.execute("SELECT device_name FROM dll_device_basic_data WHERE device_imei = %s", (device_imei,))
                    device_info = cursor.fetchone()

                    if not device_info:
                        continue

                    device_name = device_info[0]

                    # Fetch Geozone Triggered Events for the Device
                    cursor.execute("""
                        SELECT start_time, duration_of_state, date_logged, end_time, start_location, 
                            end_location, start_location_cords, end_location_cords 
                        FROM dll_state_durations 
                        WHERE device_imei=%s 
                        AND state=%s 
                        AND end_time != 'Incoming' 
                        AND date_logged BETWEEN %s AND %s
                    """, (device_imei, _ReportState, _StartDate, _EndDate))

                    state_data = cursor.fetchall()
                    if not state_data:
                        continue

                    # Prepare Data for PDF Export
                    state_report_data = []
                    for state_info in state_data:
                        _state_time = state_info[0]
                        _state_duration = int(state_info[1])
                        _date_logged = state_info[2]
                        _state_endTime = state_info[3]
                        _state_start_location = state_info[4]
                        _state_end_location = state_info[5]
                        _state_start_cords = state_info[6]
                        _state_end_cords = state_info[7]

                        # Calculate Duration Spent
                        if _state_duration < 60:
                            _DurationSpent = f"{_state_duration} Minutes"
                        else:
                            hours = _state_duration // 60
                            minutes = _state_duration % 60
                            _DurationSpent = f"{hours}Hr{'s' if hours > 1 else ''}-{minutes}Min{'s' if minutes != 1 else ''}"

                        # Append to report data
                        state_report_data.append([
                            _date_logged, _state_time, _state_endTime, 
                            _DurationSpent, _state_start_location, 
                            _state_end_location, _state_start_cords, _state_end_cords
                        ])

                    # Add data to PDF structure
                    sheet_name = f"{device_name[:15]}_{device_imei[-4:]}"
                    PDF_Data[sheet_name] = state_report_data

                # Check if there is data to export
                if not PDF_Data:
                    cursor.execute("UPDATE dll_reports_downloadable_files SET request_status=%s WHERE request_uid=%s", ('failed', _OriginRequestID,))
                    return reply("error", 400, "No State Data to Report", "")

                # Create PDF File
                reports_dir = current_app.config.get('REPORTS_DIR', 'reports-cdn')
                os.makedirs(reports_dir, exist_ok=True)
                random_suffix = random.randint(100000, 999999)
                file_name = f"{_ReportState}_state_report_{random_suffix}.pdf"
                file_path = os.path.join(reports_dir, file_name)

                # HTML Template for PDF using WeasyPrint
                html_template = Template("""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            @page {
                                size: A4 landscape;
                                margin: 1cm;
                            }
                            body {
                                font-family: Arial, sans-serif;
                                font-size: 9px;
                            }
                            h1 {
                                text-align: center;
                            }
                            h2 {
                                text-align: left;
                            }
                            .table-container {
                                max-width: 100%;
                                overflow-x: auto;
                                margin-bottom: 20px;
                            }
                            table {
                                width: 100%;
                                border-collapse: collapse;
                                font-size: 9px;
                            }
                            th, td {
                                padding: 6px;
                                border: 1px solid #ddd;
                                text-align: left;
                            }
                            th {
                                background-color: #f2f2f2;
                            }
                            tr:nth-child(even) {
                                background-color: #f9f9f9;
                            }
                        </style>
                    </head>
                    <body>
                        <h1>{{ _ReportState }} State Report</h1>
                        <br><br>
                        <h2>Report Period: {{StartDate}} To {{EndDate}}</h2>
                        <br><br>
                        {% for sheet_name, events in data.items() %}
                            <h2>Report For: {{ sheet_name }}</h2>
                            <div class="table-container">
                                <table>
                                    <thead>
                                        <tr>
                                            <th>Date Logged</th>
                                            <th>Start Time</th>
                                            <th>End Time</th>
                                            <th>Duration</th>
                                            <th>Start Location</th>
                                            <th>End Location</th>
                                            <th>Starting GPS</th>
                                            <th>End GPS</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for event in events %}
                                            <tr>
                                                {% for field in event %}
                                                    <td>{{ field }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        {% endfor %}
                    </body>
                    </html>
                """)

                # Render HTML with data
                rendered_html = html_template.render(data=PDF_Data, _ReportState=_ReportState, StartDate=_StartDate, EndDate=_EndDate)

                # Generate PDF using WeasyPrint
                HTML(string=rendered_html).write_pdf(file_path)

                # Update file path in the database
                file_url = current_app.config['base_url'] + f"{reports_dir}/{file_name}"
                cursor.execute("""
                    UPDATE dll_reports_downloadable_files
                    SET file_path = %s, request_status = 'completed'
                    WHERE request_uid = %s
                """, (file_url, _OriginRequestID))

                return reply("success", 200, f"{_ReportState} State report generated successfully", {"file_url": file_url})

    except Exception as error:
        return reply("error", 500, str(error), "")


#Statistics
# @data_handler_bp.route("/statistics/<parent_org_uid>/<account_type>/dashboard", methods=["GET"])
# def DashboardStatistics(parent_org_uid, account_type):

#     try:
#         _OrgnazationID = str(parent_org_uid)
#         _AccessType = str(account_type)

#         if(len(_OrgnazationID) > 5) and (len(_AccessType) > 2):
            
#             if(_AccessType == 'inhouse'):

            


#         else:
#             return reply("error", 400, "Some data is missing", "")

#     except Exception as error:
#         return reply("error", 500, str(error), "")
